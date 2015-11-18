from ftw.redirector import _
from ftw.redirector.interfaces import IRedirectConfig
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook
from plone import api
from zope.i18n import translate


RULES_START_ROW = 2


def create_rules_excel():
    portal = api.portal.get()
    request = portal.REQUEST

    rules = IRedirectConfig(portal).rules
    if not rules:
        rules = []

    book = Workbook()
    sheet = book.active

    title = translate(_(u'Redirect Configuration'), context=request)
    source_title = translate(_(u'label_source_path',
                             default=u'Source Path'), context=request)
    destination_title = translate(_(u'label_destination',
                                  default=u'Destination'), context=request)

    # HEADER
    sheet.title = title

    bold = Font(bold=True)
    cell = sheet.cell(column=1, row=1)
    cell.font = bold
    cell.value = source_title

    cell = sheet.cell(column=2, row=1)
    cell.font = bold
    cell.value = destination_title

    # DATA
    for rule_nr, rule in enumerate(rules):
        cell = sheet.cell(column=1, row=RULES_START_ROW+rule_nr)
        cell.value = rule['source_path']
        cell = sheet.cell(column=2, row=RULES_START_ROW+rule_nr)
        cell.value = rule['destination']

    # match the width to the longest entry
    for column in sheet.columns:
        maxwidth = 0
        for cell in column:
            if not cell.value:
                continue
            cwidth = len(cell.value)
            maxwidth = cwidth if cwidth > maxwidth else maxwidth
        # a bit more space for readability
        sheet.column_dimensions[cell.column].width = maxwidth + 5

    return save_virtual_workbook(book)


def load_rules_from_excel(excel_file):
    book = load_workbook(excel_file, read_only=True)
    sheet = book.active
    rows = sheet.rows

    # skip header rows
    for i in range(RULES_START_ROW-1):
        next(rows)

    rules = []
    for row in rows:
        source_path = row[0].value or ''
        destination = row[1].value or ''

        # ignore empty lines
        if len(source_path) == 0 and len(destination) == 0:
            continue

        rules.append({
            'source_path': source_path,
            'destination': destination
        })

    return rules
